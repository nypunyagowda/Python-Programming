import openpyxl as op
# load the workbook
wb = op.load_workbook('.\VideoSales.xlsx')
# print(type(wb)) # <class 'openpyxl.workbook.workbook.Workbook'>

# print(wb) # <openpyxl.workbook.workbook.Workbook object at 0x0000021DA9824F10>

# Call the active worksheet
ws = wb.active
# print(type(ws))

# Call the sheet by it's name
ws = wb['SalesData']
# print(ws) # <Worksheet "SalesData">

# Count the number of rows and columns that are filled in the worksheet
# print(f'Total number of rows is {ws.max_row} and total number of columns is {ws.max_column}')

# Read the data from a cell
# print(f"The value stored in cell D5 is: {ws['D5'].value}")

# Read data from multiple cells
values = [
    ws.cell(row = 1, column = i).value
    for i in range(1, ws.max_column + 1)
]
# print(values)

data = [
    ws.cell(row = 1, column = i).value
    for i in range(1, ws.max_column + 1)
]
# print(values)

# Reading data from a range of cells
list1 = list()
for value in ws.iter_rows(min_row = 1, max_row = 5, min_col = 1, max_col = 4, values_only = True):
    list1.append(value)
# print(list1)

# Display the values in a tabular method
for e1,e2,e3,e4 in list1:
    '''unpacking of a tuple'''
    # (print("{:<10}{:<35}{:<10}{:<5}".format(e1,e2,e3,e4)))
    '''width of 1st column is 10, 2nd column is 35, 3rd column is 10, 4th column is 5'''

# Write to a cell
ws['K1'] = 'Total Sales'

# Add a new row
# new_row = (
#     31, 'Cricket Premier League', 'PC', 2025, 'Sports',
#     'GameStudio', 2.60, 1.40, 3.80, 2.20, 10.0
# )
# ws.append(new_row)
# wb.save('VideoSales.xlsx')
# print(ws.max_row)

# Delete a row
# ws.delete_rows(ws.max_row, 1)
# '''First parameter - row number from where the deletion starts.
#    Second parameter - How many rows are to be deleted'''
# print(ws.max_row)
# wb.save('VideoSales.xlsx')
# print(ws.max_row)

# ------------ Excel Formulas ----------------

# Average
ws['M1'] = 'Average Sales'
ws['M2'] = 'AVERAGE(K2:K31)'
wb.save('VideoSales.xlsx')

# CountA
'''It counts the number of cells that are populated in a given range'''
# new_row = (
#     29, 'Cricket Premier League', 'PC', ' ', 'Sports',
#     'GameStudio', 2.60, 1.40, 3.80, 2.20, 10.0
# )
# ws.append(new_row)
# wb.save('VideoSales.xlsx')
# print(ws.max_row)

# ws['N1'] = 'Number of rows that have the value'
# ws['N2'] = '=COUNTA(D1:D32)'
# wb.save('VideoSales.xlsx')

# Countif
'''This counts the number of cells that meet a criteria'''
# ws['01'] = 'Number of rows with Sports Genre'
# ws['02'] = 'COUNTIF(E2:E32, "Sports")'
# wb.save('VideoSales.xlsx')

# Sumif
ws['M5'] = 'Total Sports Sales'
ws['M6'] = '=SUMIF(E2:E32, "Sports", K2:K32)'
'''First parameter - where should the criteria be checked
   Second parameter - what is the criteria
   Third parameter - what should be summed'''
wb.save('VideoSales.xlsx')
