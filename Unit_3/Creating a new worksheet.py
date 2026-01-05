import openpyxl as op

wb = op.load_workbook('.\VideoSales.xlsx')
ws = wb.active
# ws = wb['SalesData']
# wb.create_sheet('New Sheet')
# wb.save('VideoSales.xlsx')
# print(wb.sheetnames)

# --------- Title of active sheet -------------
print(ws.title)

# --------- Rename a sheet ----------------
# ws = wb['New Sheet1']
# print(ws.title)
ws.title = 'A New Sheet'
# wb.save('VideoSales.xlsx')
# print(wb.sheetnames)
print(wb.sheetnames) 

# Duplicate a worksheet
new_ws = wb.copy_worksheet(wb['A New Sheet1'])
new_ws.title = 'A New Sheet'
wb.save('VideoSales.xlsx')
print(wb.sheetnames)