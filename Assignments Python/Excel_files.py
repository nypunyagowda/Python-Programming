from openpyxl import Workbook, load_workbook
import os

FILENAME = "records.xlsx"

def create_file():
    wb = Workbook()
    sheet = wb.active
    sheet.append(["ID", "Name", "Marks"])
    wb.save(FILENAME)
    print("Excel file created.")

def add_record():
    wb = load_workbook(FILENAME)
    sheet = wb.active

    id = int(input("Enter ID: "))
    name = input("Enter Name: ")
    marks = int(input("Enter Marks: "))

    sheet.append([id, name, marks])
    wb.save(FILENAME)
    print("Record added.")

def view_records():
    wb = load_workbook(FILENAME)
    sheet = wb.active

    print("\nID   Name    Marks")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row[0], row[1], row[2])

def update_marks():
    wb = load_workbook(FILENAME)
    sheet = wb.active

    search_id = int(input("Enter ID to update: "))

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == search_id:
            new_marks = int(input("Enter new marks: "))
            row[2].value = new_marks
            wb.save(FILENAME)
            print("Marks updated.")
            return

    print("ID not found.")

# Main Program
if not os.path.exists(FILENAME):
    create_file()

while True:
    print("\n1. Add Record")
    print("2. View Records")
    print("3. Update Marks")
    print("4. Exit")

    choice = input("Enter choice: ")

    if choice == "1":
        add_record()
    elif choice == "2":
        view_records()
    elif choice == "3":
        update_marks()
    elif choice == "4":
        break
    else:
        print("Invalid choice.")
